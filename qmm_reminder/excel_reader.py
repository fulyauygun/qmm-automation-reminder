"""Read the QMM control list Excel file into Document objects.

The file is opened read-only in place (e.g. directly from the network
share); nothing is copied or written back.
"""

from __future__ import annotations

import hashlib
import logging
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .config import ExcelConfig
from .models import Document

log = logging.getLogger(__name__)

_DATE_FORMATS = ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d")
_HEADER_SCAN_ROWS = 15


class ExcelReadError(Exception):
    pass


def _normalize(value) -> str:
    """Normalize a header/id cell: collapse whitespace, casefold."""
    return " ".join(str(value).split()).casefold()


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def add_years(d: date, years: int) -> date:
    """d plus N years; 29 Feb maps to 28 Feb in non-leap years."""
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        return d.replace(year=d.year + years, day=28)


def _find_header_row(ws, wanted: dict[str, str]) -> tuple[int, dict[str, int]]:
    """Locate the header row and map logical column name -> 0-based index.

    Cells are addressed by tuple position, not ``cell.column``: in
    read-only mode openpyxl yields bare ``EmptyCell`` objects that carry
    no coordinates.
    """
    wanted_norm = {logical: _normalize(text) for logical, text in wanted.items()}
    best: tuple[int, dict[str, int]] | None = None
    for row_no, row in enumerate(
        ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_ROWS), start=1
    ):
        headers = {
            _normalize(cell.value): pos
            for pos, cell in enumerate(row)
            if cell.value not in (None, "")
        }
        mapping = {
            logical: headers[text]
            for logical, text in wanted_norm.items()
            if text in headers
        }
        if best is None or len(mapping) > len(best[1]):
            best = (row_no, mapping)
    if best is None or not all(
        logical in best[1] for logical in ExcelConfig.REQUIRED_COLUMNS
    ):
        found = sorted(best[1]) if best else []
        raise ExcelReadError(
            "Could not locate the header row: required columns "
            f"{list(ExcelConfig.REQUIRED_COLUMNS)} not all found "
            f"(matched: {found}). Check excel.columns in config.yaml."
        )
    return best


def read_documents(
    cfg: ExcelConfig, validity_years: int
) -> tuple[list[Document], list[str]]:
    """Return (documents, warnings). Rows that cannot be interpreted are
    skipped and reported as warnings so the run never fails silently."""
    path = Path(cfg.path)
    if not path.is_file():
        raise ExcelReadError(f"Excel file not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[cfg.sheet] if cfg.sheet else wb.active
        if ws is None:
            raise ExcelReadError(f"Worksheet not found: {cfg.sheet!r}")
        header_row, col_map = _find_header_row(ws, cfg.columns)

        documents: list[Document] = []
        warnings: list[str] = []
        for row_no, row in enumerate(
            ws.iter_rows(min_row=header_row + 1), start=header_row + 1
        ):
            if not row:
                continue
            values = {
                logical: (row[pos].value if pos < len(row) else None)
                for logical, pos in col_map.items()
            }
            if all(v in (None, "") for v in values.values()):
                continue  # blank spacer row

            revision_date = _parse_date(values.get("revision_date"))
            due = _parse_date(values.get("due_date"))
            if due is None and revision_date is not None:
                due = add_years(revision_date, validity_years)
            if due is None:
                warnings.append(
                    f"Row {row_no}: no usable due date or revision date - skipped"
                )
                continue

            id_parts = [_normalize(values.get(c) or "") for c in cfg.id_columns]
            if not any(id_parts):
                warnings.append(
                    f"Row {row_no}: identity columns {cfg.id_columns} all empty"
                    " - skipped"
                )
                continue
            doc_id = hashlib.sha1("|".join(id_parts).encode("utf-8")).hexdigest()[:16]

            title = str(
                values.get("title")
                or values.get("approval_doc")
                or values.get("section")
                or f"Satır {row_no}"
            ).strip()

            documents.append(
                Document(
                    doc_id=doc_id,
                    title=title,
                    section=str(values.get("section") or "").strip(),
                    revision_no=str(values.get("revision_no") or "").strip(),
                    prepared_by=str(values.get("prepared_by") or "").strip(),
                    revision_date=revision_date,
                    due_date=due,
                    source_row=row_no,
                )
            )
        return documents, warnings
    finally:
        wb.close()


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_EMAIL_HEADERS = {"e-posta", "eposta", "e-mail", "email", "mail"}
_NAME_HEADERS = {"ad", "ad soyad", "isim", "name"}
_ACTIVE_HEADERS = {"aktif", "active"}
_INACTIVE_VALUES = {"hayır", "hayir", "no", "0", "false", "pasif"}


def read_recipients(
    path: Path, sheet_name: str
) -> tuple[list[tuple[str, str]], list[str]]:
    """Read notification recipients from a worksheet.

    The sheet is maintained by the QMM team directly in Excel: one row per
    person, an "E-posta" column (required), an optional "Ad" column (used
    for Teams mentions) and an optional "Aktif" column ("Hayır" disables a
    row without deleting it). Returns ((name, e-mail) pairs, warnings);
    a missing sheet is a warning, not a crash, so a broken recipients list
    never stops the notifications.
    """
    if not path.is_file():
        return [], [f"Recipients: Excel file not found: {path}"]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        actual = next(
            (n for n in wb.sheetnames if _normalize(n) == _normalize(sheet_name)),
            None,
        )
        if actual is None:
            return [], [
                f"Recipients sheet {sheet_name!r} not found in {path.name}"
                f" (sheets: {wb.sheetnames})"
            ]
        ws = wb[actual]

        email_col = name_col = active_col = header_row = None
        for row_no, row in enumerate(
            ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_ROWS), start=1
        ):
            for pos, cell in enumerate(row):
                header = _normalize(cell.value) if cell.value else ""
                if header in _EMAIL_HEADERS:
                    email_col, header_row = pos, row_no
                elif header in _NAME_HEADERS:
                    name_col = pos
                elif header in _ACTIVE_HEADERS:
                    active_col = pos
            if email_col is not None:
                break
        if email_col is None:
            return [], [
                f"Recipients sheet {actual!r}: no 'E-posta' header found"
            ]

        def _cell(row, pos):
            return row[pos].value if pos is not None and pos < len(row) else None

        people: list[tuple[str, str]] = []
        warnings: list[str] = []
        for row_no, row in enumerate(
            ws.iter_rows(min_row=header_row + 1), start=header_row + 1
        ):
            value = _cell(row, email_col)
            if value in (None, ""):
                continue
            flag = _normalize(_cell(row, active_col) or "")
            if flag in _INACTIVE_VALUES:
                continue
            email = str(value).strip()
            if not _EMAIL_RE.match(email):
                warnings.append(
                    f"Recipients sheet row {row_no}: invalid address"
                    f" {email!r} - skipped"
                )
                continue
            name = str(_cell(row, name_col) or "").strip() or email.split("@")[0]
            if email.casefold() not in {e.casefold() for _, e in people}:
                people.append((name, email))
        return people, warnings
    finally:
        wb.close()
