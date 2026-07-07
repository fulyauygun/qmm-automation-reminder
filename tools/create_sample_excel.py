"""Generate a fictitious control list for testing / demonstration.

Usage:  py -3 tools/create_sample_excel.py [output.xlsx]

The rows are placed relative to today so every reminder scenario
(no action, T-30, T-15, T-7, T-1, due today, overdue) is exercised.
All names and contents are invented - no real company data.
"""

from __future__ import annotations

import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

HEADERS = [
    "Bölüm",
    "Revizyon No",
    "Revizyon Tarihi",
    "Hazırlayan",
    "Revizyon İçeriği",
    "Değişiklik Olmaması Durumunda Güncellenmesi Gereken İlk Tarih",
    "Onay Dokümanı",
]


def build(path: Path, today: date | None = None) -> None:
    today = today or date.today()

    def rev(due_in_days: int) -> tuple[date, date]:
        due = today + timedelta(days=due_in_days)
        return due.replace(year=due.year - 3), due

    rows = []
    scenarios = [
        ("Kalite Güvence", "QMM-TL-001", 45),   # no reminder yet
        ("Üretim", "QMM-TL-002", 30),           # T-30
        ("Üretim", "QMM-TL-003", 14),           # T-15 (catch-up)
        ("Lojistik", "QMM-TL-004", 7),          # T-7
        ("Kalite Kontrol", "QMM-TL-005", 1),    # T-1
        ("Bakım", "QMM-TL-006", 0),             # due today
        ("Kalite Güvence", "QMM-TL-007", -10),  # overdue
        ("Ar-Ge", "QMM-TL-008", 200),           # far in the future
    ]
    for i, (section, doc_no, due_in) in enumerate(scenarios, start=1):
        rev_date, due = rev(due_in)
        rows.append([
            section,
            f"{i % 4:02d}",
            rev_date,
            "A. Yılmaz" if i % 2 else "B. Demir",
            "Örnek revizyon açıklaması (kurgusal)",
            due,
            f"{doc_no}-ONAY",
        ])
    # one row with an empty due-date cell -> falls back to revision date + 3y
    rows.append([
        "İnsan Kaynakları", "01",
        today.replace(year=today.year - 3) + timedelta(days=20),
        "C. Kaya", "Örnek revizyon açıklaması (kurgusal)", None,
        "QMM-TL-009-ONAY",
    ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Kontrol Listesi"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    for col, width in zip("ABCDEFG", (18, 12, 14, 14, 34, 20, 18)):
        ws.column_dimensions[col].width = width
    for r in range(2, len(rows) + 2):
        ws.cell(row=r, column=3).number_format = "DD.MM.YYYY"
        ws.cell(row=r, column=6).number_format = "DD.MM.YYYY"

    # Self-service e-mail recipient sheet maintained by the QMM team.
    rcp = wb.create_sheet("Bildirim Alıcıları")
    rcp.append(["Ad", "E-posta", "Aktif"])
    rcp.append(["Süreç Lideri (kurgusal)", "surec.lideri@example.com", "Evet"])
    rcp.append(["QMM Ekibi (kurgusal)", "qmm.ekibi@example.com", ""])
    rcp.append(["Eski Çalışan (kurgusal)", "ayrilan.kisi@example.com", "Hayır"])
    for col, width in zip("ABC", (26, 30, 8)):
        rcp.column_dimensions[col].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Sample file written: {path}")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "sample_data/QMM Working Instructions Control List.xlsx"
    )
    build(out)
